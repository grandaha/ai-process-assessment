from pathlib import Path

import openpyxl

from engine.model import load_inputs
from engine.run import build_results
from engine.workbook import write_workbook

FIXTURE = Path(__file__).parent / "fixtures" / "lattice" / "model"


def test_workbook_has_expected_tabs_and_formulas(tmp_path):
    inp = load_inputs(FIXTURE)
    res = build_results(FIXTURE)
    out = tmp_path / "financial-model.xlsx"
    write_workbook(inp, res, out)

    wb = openpyxl.load_workbook(out)  # data_only=False -> formula strings
    for tab in ("Inputs", "Value (P5)", "Scores (P6)", "Costs (P8.5)",
                "Business Case (P9)", "Wave-1 Aggregate"):
        assert tab in wb.sheetnames

    # Downstream cells must be formulas (start with "="), not literals.
    value_tab = wb["Value (P5)"]
    formula_cells = [c.value for row in value_tab.iter_rows() for c in row
                     if isinstance(c.value, str) and c.value.startswith("=")]
    assert formula_cells, "Value tab must contain at least one =formula"


def test_workbook_sets_full_calc_on_load(tmp_path):
    inp = load_inputs(FIXTURE)
    res = build_results(FIXTURE)
    out = tmp_path / "wb.xlsx"
    write_workbook(inp, res, out)
    wb = openpyxl.load_workbook(out)
    assert wb.calculation.fullCalcOnLoad is True


def test_inputs_tab_writes_resolved_volume(tmp_path):
    inputs = load_inputs(FIXTURE)
    results = build_results(FIXTURE)
    out = tmp_path / "wb.xlsx"
    write_workbook(inputs, results, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Inputs"]
    # Header row 1; OPP-001 is row 2; volume is column E (index 5).
    assert ws.cell(row=2, column=1).value == "OPP-001"
    assert ws.cell(row=2, column=5).value == 8000  # resolved PROC-01 volume * 1.0
