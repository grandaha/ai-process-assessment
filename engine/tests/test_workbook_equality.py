import json
import shutil
from pathlib import Path

import openpyxl
import pytest

from engine.model import load_inputs
from engine.run import build_results
from engine.workbook import write_workbook

FIXTURE = Path(__file__).parent / "fixtures" / "lattice" / "model"
formulas = pytest.importorskip("formulas")


def test_workbook_results_equal_results_json(tmp_path):
    inp = load_inputs(FIXTURE)
    res = build_results(FIXTURE)
    out = tmp_path / "financial-model.xlsx"
    write_workbook(inp, res, out)

    xl = formulas.ExcelModel().loads(str(out)).finish()
    sol = xl.calculate()

    def cell(sheet, a1):
        # formulas emits keys like "'[fm.xlsx]WAVE-1 AGGREGATE'!B2"; match on the
        # sheet-name + cell suffix (sheet name may be prefixed by [workbook]).
        for k, v in sol.items():
            ku = k.upper()
            if ku.endswith(f"{sheet.upper()}'!{a1}") or ku.endswith(f"{sheet.upper()}!{a1}"):
                return float(v.value[0, 0])
        raise KeyError(f"{sheet}!{a1}")

    # Wave-1 aggregate investment low/high must equal results.json.
    inv = res["wave1_aggregate"]["investment"]
    assert cell("Wave-1 Aggregate", "B2") == pytest.approx(inv["low"])
    assert cell("Wave-1 Aggregate", "C2") == pytest.approx(inv["high"])
    av = res["wave1_aggregate"]["value"]
    assert cell("Wave-1 Aggregate", "B3") == pytest.approx(av["low"])
    assert cell("Wave-1 Aggregate", "C3") == pytest.approx(av["high"])
    # Workbook ROUNDs payback to 2dp (like compute.payback), so it equals
    # results.json exactly — no tolerance fudge needed.
    pb = res["wave1_aggregate"]["payback_years"]
    assert cell("Wave-1 Aggregate", "B4") == pytest.approx(pb["low"])
    assert cell("Wave-1 Aggregate", "C4") == pytest.approx(pb["high"])


def test_workbook_aggregate_matches_results_when_one_initiative_pending(tmp_path):
    # Break OPP-002's labor_hours -> its cost is PENDING -> excluded from aggregate.
    model = tmp_path / "model"
    shutil.copytree(FIXTURE, model)
    costs = json.loads((model / "costs.json").read_text())
    for c in costs:
        if c["opp_id"] == "OPP-002":
            c["labor_hours"] = None
    (model / "costs.json").write_text(json.dumps(costs))

    inp = load_inputs(model)
    res = build_results(model)
    # results.json must exclude OPP-002 from the investment aggregate.
    assert res["costs"]["OPP-002"]["rom"] == "PENDING"
    inv = res["wave1_aggregate"]["investment"]
    assert inv == {"low": 155_250.0, "high": 465_750.0}  # OPP-001 only

    out = tmp_path / "financial-model.xlsx"
    write_workbook(inp, res, out)
    xl = formulas.ExcelModel().loads(str(out)).finish()
    sol = xl.calculate()

    def cell(sheet, a1):
        for k, v in sol.items():
            ku = k.upper()
            if ku.endswith(f"{sheet.upper()}'!{a1}") or ku.endswith(f"{sheet.upper()}!{a1}"):
                return float(v.value[0, 0])
        raise KeyError(f"{sheet}!{a1}")

    # Workbook aggregate must equal results.json (OPP-002 PENDING -> ignored by SUM).
    assert cell("Wave-1 Aggregate", "B2") == pytest.approx(inv["low"])
    assert cell("Wave-1 Aggregate", "C2") == pytest.approx(inv["high"])


def test_workbook_scores_round_to_match_results(tmp_path):
    # A composite that is not a clean multiple of 6 exposes rounding drift:
    # [4,4,4,4,4,5] -> 25/6 = 4.1666...  results.json rounds to 4.17; the
    # workbook formula must ROUND too or the two sources diverge.
    model = tmp_path / "model"
    shutil.copytree(FIXTURE, model)
    scores = json.loads((model / "scores.json").read_text())
    for s in scores:
        if s["opp_id"] == "OPP-001":
            s["dimensions"] = [4, 4, 4, 4, 4, 5]
    (model / "scores.json").write_text(json.dumps(scores))

    inp = load_inputs(model)
    res = build_results(model)
    assert res["scores"]["OPP-001"] == 4.17  # rounded in the engine

    out = tmp_path / "financial-model.xlsx"
    write_workbook(inp, res, out)
    xl = formulas.ExcelModel().loads(str(out)).finish()
    sol = xl.calculate()

    def cell(sheet, a1):
        for k, v in sol.items():
            ku = k.upper()
            if ku.endswith(f"{sheet.upper()}'!{a1}") or ku.endswith(f"{sheet.upper()}!{a1}"):
                return float(v.value[0, 0])
        raise KeyError(f"{sheet}!{a1}")

    # Workbook composite (row 2 = OPP-001) must equal the rounded results value.
    assert cell("Scores (P6)", "B2") == pytest.approx(res["scores"]["OPP-001"])


def test_workbook_aggregate_renders_pending_when_all_initiatives_pending(tmp_path):
    # Every initiative missing both cost labor and value inputs -> every aggregate
    # is PENDING in results.json. The workbook must render literal "PENDING" (not a
    # formula that divides by a zero SUM -> #DIV/0!), matching results.json.
    model = tmp_path / "model"
    shutil.copytree(FIXTURE, model)
    costs = json.loads((model / "costs.json").read_text())
    for c in costs:
        c["labor_hours"] = None  # -> cost PENDING for every initiative
    (model / "costs.json").write_text(json.dumps(costs))
    values = json.loads((model / "value.json").read_text())
    for v in values:
        v["improvement_low"] = None  # -> value PENDING for every initiative
    (model / "value.json").write_text(json.dumps(values))

    inp = load_inputs(model)
    res = build_results(model)
    assert res["wave1_aggregate"]["investment"] == "PENDING"
    assert res["wave1_aggregate"]["value"] == "PENDING"
    assert res["wave1_aggregate"]["payback_years"] == "PENDING"

    out = tmp_path / "financial-model.xlsx"
    write_workbook(inp, res, out)

    # Read literal cells with openpyxl (these are text, not formulas).
    wb = openpyxl.load_workbook(out)
    ws = wb["Wave-1 Aggregate"]
    assert ws["B2"].value == "PENDING" and ws["C2"].value == "PENDING"  # investment
    assert ws["B3"].value == "PENDING" and ws["C3"].value == "PENDING"  # annual_value
    assert ws["B4"].value == "PENDING" and ws["C4"].value == "PENDING"  # payback


def test_workbook_equals_results_exactly_with_fractional_inputs(tmp_path):
    # Inputs chosen so the cost chain produces sub-cent intermediates (labor_rate
    # 200.33, improvement 0.55/0.67), exercising rounding parity. Every workbook
    # cell must equal results.json EXACTLY — no tolerance — because compute and
    # the workbook now both ROUND at each step from rounded predecessors.
    model = tmp_path / "model"
    shutil.copytree(FIXTURE, model)
    costs = json.loads((model / "costs.json").read_text())
    for c in costs:
        if c["opp_id"] == "OPP-001":
            c["labor_rate"] = 200.33
            c["labor_hours"] = 801
    (model / "costs.json").write_text(json.dumps(costs))
    values = json.loads((model / "value.json").read_text())
    for v in values:
        if v["opp_id"] == "OPP-001":
            v["improvement_low"], v["improvement_high"] = 0.55, 0.67
            v["volume"], v["rate"] = 7333, 101
    (model / "value.json").write_text(json.dumps(values))

    inp = load_inputs(model)
    res = build_results(model)
    out = tmp_path / "financial-model.xlsx"
    write_workbook(inp, res, out)
    sol = formulas.ExcelModel().loads(str(out)).finish().calculate()

    def cell(sheet, a1):
        for k, v in sol.items():
            ku = k.upper()
            if ku.endswith(f"{sheet.upper()}'!{a1}") or ku.endswith(f"{sheet.upper()}!{a1}"):
                return float(v.value[0, 0])
        raise KeyError(f"{sheet}!{a1}")

    # "Exactly" = same rounded values; pytest.approx (default rel 1e-6) only
    # absorbs the float-serialization noise of writing/reading the .xlsx, far
    # tighter than the <0.01 drift this PR removes.
    # OPP-001 is workbook row 2 across all per-initiative tabs.
    cb = res["costs"]["OPP-001"]
    assert cell("Costs (P8.5)", "B2") == pytest.approx(cb["labor"])
    assert cell("Costs (P8.5)", "C2") == pytest.approx(cb["change_mgmt"])
    assert cell("Costs (P8.5)", "D2") == pytest.approx(cb["subtotal"])
    assert cell("Costs (P8.5)", "E2") == pytest.approx(cb["contingency"])
    assert cell("Costs (P8.5)", "F2") == pytest.approx(cb["total"])
    assert cell("Costs (P8.5)", "G2") == pytest.approx(cb["rom"]["low"])
    assert cell("Costs (P8.5)", "H2") == pytest.approx(cb["rom"]["high"])

    vr = res["value"]["OPP-001"]
    assert cell("Value (P5)", "B2") == pytest.approx(vr["low"])
    assert cell("Value (P5)", "C2") == pytest.approx(vr["high"])

    agg = res["wave1_aggregate"]
    assert cell("Wave-1 Aggregate", "B2") == pytest.approx(agg["investment"]["low"])
    assert cell("Wave-1 Aggregate", "C2") == pytest.approx(agg["investment"]["high"])
    assert cell("Wave-1 Aggregate", "B3") == pytest.approx(agg["value"]["low"])
    assert cell("Wave-1 Aggregate", "C3") == pytest.approx(agg["value"]["high"])
    assert cell("Wave-1 Aggregate", "B4") == pytest.approx(agg["payback_years"]["low"])
    assert cell("Wave-1 Aggregate", "C4") == pytest.approx(agg["payback_years"]["high"])
