"""openpyxl writer: inputs as literal cells, every downstream value as a live =formula.

Cross-app safe: plain arithmetic, SUM, and A1/sheet references only. No Excel-only
functions, named ranges, or tables. fullCalcOnLoad=True so Numbers/Sheets recompute.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl


def _q(sheet_name: str) -> str:
    # Sheet names with spaces/parens must be single-quoted in references.
    return f"'{sheet_name}'"


def write_workbook(inputs, results, out_path) -> Path:
    out_path = Path(out_path)
    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True

    ws_in = wb.active
    ws_in.title = "Inputs"
    wave1 = [i for i in inputs.initiatives if i.wave == 1]

    # --- Inputs tab: literal cells. One row per Wave-1 initiative. ---
    ws_in.append(["OPP-ID", "Name",
                  "impr_low", "impr_high", "volume", "rate",
                  "labor_hours", "labor_rate", "tech_cost", "integration_cost",
                  "change_mgmt_pct", "contingency_pct",
                  "d1", "d2", "d3", "d4", "d5", "d6"])
    row_of = {}
    for idx, init in enumerate(wave1, start=2):
        opp = init.opp_id
        v = inputs.value.get(opp)
        c = inputs.costs.get(opp)
        s = inputs.scores.get(opp)
        ws_in.append([
            opp, init.name,
            getattr(v, "improvement_low", None), getattr(v, "improvement_high", None),
            getattr(v, "volume", None), getattr(v, "rate", None),
            getattr(c, "labor_hours", None), getattr(c, "labor_rate", None),
            getattr(c, "tech_cost", None), getattr(c, "integration_cost", None),
            getattr(c, "change_mgmt_pct", None), getattr(c, "contingency_pct", None),
            *(s.dimensions if s else [None] * 6),
        ])
        row_of[opp] = idx

    IN = _q("Inputs")

    # --- PENDING predicates (results.json is the source of truth) ---
    # Missing inputs render as PENDING in both results.json AND the workbook —
    # never as a fabricated number (spec §12). For a PENDING initiative we write
    # the literal text "PENDING" into its downstream cells; SUM ignores text, so
    # aggregates sum over present members only, matching results.json.
    def _cost_pending(opp) -> bool:
        return results["costs"].get(opp, {}).get("rom") == "PENDING"

    def _value_pending(opp) -> bool:
        return results["value"].get(opp, "PENDING") == "PENDING"

    def _score_pending(opp) -> bool:
        return results["scores"].get(opp, "PENDING") == "PENDING"

    # --- Value (P5): improvement × volume × rate ---
    ws_v = wb.create_sheet("Value (P5)")
    ws_v.append(["OPP-ID", "value_low", "value_high"])
    for opp, r in row_of.items():
        if _value_pending(opp):
            ws_v.append([opp, "PENDING", "PENDING"])
        else:
            ws_v.append([opp,
                         f"={IN}!C{r}*{IN}!E{r}*{IN}!F{r}",   # impr_low*volume*rate
                         f"={IN}!D{r}*{IN}!E{r}*{IN}!F{r}"])  # impr_high*volume*rate

    # --- Scores (P6): mean of 6 dims ---
    ws_s = wb.create_sheet("Scores (P6)")
    ws_s.append(["OPP-ID", "composite"])
    for opp, r in row_of.items():
        if _score_pending(opp):
            ws_s.append([opp, "PENDING"])
        else:
            # ROUND to 2dp to match compute.score_composite (round(sum/6, 2)).
            # ROUND is common to Excel / Google Sheets / Apple Numbers.
            ws_s.append([opp, f"=ROUND(SUM({IN}!M{r}:R{r})/6, 2)"])

    # --- Costs (P8.5): roll-up ---
    ws_c = wb.create_sheet("Costs (P8.5)")
    ws_c.append(["OPP-ID", "labor", "change_mgmt", "subtotal", "contingency",
                 "total", "rom_low", "rom_high"])
    cost_row = {}
    for i, (opp, r) in enumerate(row_of.items(), start=2):
        if _cost_pending(opp):
            # 7 PENDINGs: labor, change_mgmt, subtotal, contingency, total, rom_low, rom_high
            ws_c.append([opp, "PENDING", "PENDING", "PENDING", "PENDING",
                         "PENDING", "PENDING", "PENDING"])
        else:
            # labor=G*H ; cm=labor*K ; subtotal=labor+I+J+cm ; cont=subtotal*L ; total=subtotal+cont
            ws_c.append([
                opp,
                f"={IN}!G{r}*{IN}!H{r}",                              # B labor
                f"=B{i}*{IN}!K{r}",                                   # C change_mgmt
                f"=B{i}+{IN}!I{r}+{IN}!J{r}+C{i}",                    # D subtotal
                f"=D{i}*{IN}!L{r}",                                   # E contingency
                f"=D{i}+E{i}",                                        # F total
                f"=F{i}*0.5",                                         # G rom_low
                f"=F{i}*1.5",                                         # H rom_high
            ])
        cost_row[opp] = i

    # --- Business Case (P9): per-initiative ROM (references Costs tab) ---
    ws_bc = wb.create_sheet("Business Case (P9)")
    ws_bc.append(["OPP-ID", "rom_low", "rom_high", "value_low", "value_high"])
    CO = _q("Costs (P8.5)")
    VA = _q("Value (P5)")
    for i, opp in enumerate(row_of, start=2):
        cr = cost_row[opp]
        ws_bc.append([opp,
                      f"={CO}!G{cr}", f"={CO}!H{cr}",
                      f"={VA}!B{i}", f"={VA}!C{i}"])

    # --- Wave-1 Aggregate: SUMs + payback ---
    # When a whole aggregate is PENDING in results.json (every member missing),
    # write literal "PENDING" rather than a formula — matching results.json and
    # avoiding a #DIV/0! in payback when the value SUM is 0 (compute.payback
    # returns PENDING for any non-positive annual value). The payback formula is
    # therefore only emitted when results gives a real number, so B2/C3 never
    # divides by zero.
    agg = results.get("wave1_aggregate", {})
    inv_pending = agg.get("investment") == "PENDING"
    val_pending = agg.get("value") == "PENDING"
    pay_pending = agg.get("payback_years") == "PENDING"

    ws_a = wb.create_sheet("Wave-1 Aggregate")
    n = len(row_of)
    last = 1 + n  # rows 2..last in Business Case
    BC = _q("Business Case (P9)")
    ws_a.append(["metric", "low", "high"])
    ws_a.append(["investment",
                 "PENDING" if inv_pending else f"=SUM({BC}!B2:B{last})",
                 "PENDING" if inv_pending else f"=SUM({BC}!C2:C{last})"])
    ws_a.append(["annual_value",
                 "PENDING" if val_pending else f"=SUM({BC}!D2:D{last})",
                 "PENDING" if val_pending else f"=SUM({BC}!E2:E{last})"])
    # payback best = investment.low/value.high ; worst = investment.high/value.low
    ws_a.append(["payback_years",
                 "PENDING" if pay_pending else "=B2/C3",
                 "PENDING" if pay_pending else "=C2/B3"])

    wb.save(out_path)
    return out_path
